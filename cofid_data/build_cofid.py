"""
Build CoFID foods JSON + FAISS index structured identically to usda_foods.json.

CoFID does not provide the exact volumetric portion data we use for USDA-style
density calculations, so this builder preserves calorie/macronutrient data only
and leaves density fields empty unless a future exact source is added.

Output files:
  cofid_data/cofid_foods.json        — list of food dicts
  cofid_data/cofid_food_names.json   — parallel list of names (for FAISS lookup)
  cofid_data/cofid_faiss.index       — all-MiniLM-L6-v2 FAISS index
"""

import json, re, sys
import numpy as np
import openpyxl
import faiss
from pathlib import Path
from sentence_transformers import SentenceTransformer

XLSX = Path(__file__).parent / "CoFID.xlsx"
OUT_DIR = Path(__file__).parent


def safe_float(v):
    if v is None:
        return None
    s = str(v).strip()
    if s in ("N", "Tr", "None", "", "—", "n"):
        return None
    # handle trace values
    if s.lower() == "tr":
        return 0.0
    try:
        return float(s)
    except Exception:
        return None


def parse_specific_gravity_map(wb):
    ws = wb["1.2 Factors"]
    gravity_by_code = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        code = row[0]
        if not code:
            continue
        specific_gravity = safe_float(row[8])
        if specific_gravity is None:
            continue
        if 0.05 <= specific_gravity <= 3.0:
            gravity_by_code[str(code)] = round(float(specific_gravity), 3)
    return gravity_by_code


def parse_cofid():
    wb = openpyxl.load_workbook(str(XLSX), read_only=True, data_only=True)
    gravity_by_code = parse_specific_gravity_map(wb)
    ws = wb["1.3 Proximates"]

    foods = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        code = row[0]
        name = row[1]
        if not name or not code:
            continue
        name = str(name).strip()
        if not name:
            continue

        water   = safe_float(row[7])
        protein = safe_float(row[9])
        fat     = safe_float(row[10])
        carbs   = safe_float(row[11])
        kcal    = safe_float(row[12])

        if kcal is None:
            continue  # skip entries with no calorie data

        density = gravity_by_code.get(str(code))

        foods.append({
            "food_code":          str(code),
            "description":        name,
            "calories_per_100g":  round(float(kcal), 1),
            "water_g":            water,
            "protein_g":          protein,
            "fat_g":              fat,
            "carbs_g":            carbs,
            "density_g_ml":       density,
            "density_method":     "cofid_specific_gravity" if density is not None else None,
            "source":             "cofid",
        })

    print(f"Parsed {len(foods)} CoFID foods")
    return foods


def build_index(foods):
    model = SentenceTransformer("all-MiniLM-L6-v2")
    names = [f["description"] for f in foods]

    print(f"Embedding {len(names)} names...")
    vecs = model.encode(names, batch_size=256, normalize_embeddings=True,
                        show_progress_bar=True)
    vecs = np.array(vecs, dtype=np.float32)

    index = faiss.IndexFlatIP(vecs.shape[1])
    index.add(vecs)
    print(f"FAISS index: {index.ntotal} vectors, dim={vecs.shape[1]}")
    return index, names


def main():
    foods = parse_cofid()

    out_foods = OUT_DIR / "cofid_foods.json"
    out_names = OUT_DIR / "cofid_food_names.json"
    out_index = OUT_DIR / "cofid_faiss.index"

    names = [f["description"] for f in foods]
    with open(out_foods, "w") as f:
        json.dump(foods, f)
    with open(out_names, "w") as f:
        json.dump(names, f)
    print(f"Saved {out_foods}")

    index, _ = build_index(foods)
    faiss.write_index(index, str(out_index))
    print(f"Saved {out_index}")


if __name__ == "__main__":
    main()
